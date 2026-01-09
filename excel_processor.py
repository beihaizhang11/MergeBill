"""
Excel处理器 - 负责读取、提取和合并Excel数据
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os
from datetime import datetime


class ExcelProcessor:
    def __init__(self):
        pass
    
    def read_cell_value(self, file_path, cell_ref):
        """
        读取Excel文件中指定单元格的值
        
        Args:
            file_path: Excel文件路径
            cell_ref: 单元格引用，如 'A1', 'B2'
        
        Returns:
            单元格的值
        """
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            value = ws[cell_ref.upper()].value
            wb.close()
            return value
        except Exception as e:
            print(f"读取单元格 {cell_ref} 失败: {e}")
            return None
    
    def find_settlement_amount(self, ws, search_column="D", search_keyword="折后总计"):
        """
        在指定列中搜索关键词，并返回其右侧单元格的值
        
        Args:
            ws: openpyxl工作表对象
            search_column: 搜索的列字母，如"D"
            search_keyword: 搜索的关键词，如"折后总计"
        
        Returns:
            结算金额数值，如果未找到则返回None
        """
        try:
            # 计算右侧列（搜索列的下一列）
            from openpyxl.utils import column_index_from_string, get_column_letter
            search_col_idx = column_index_from_string(search_column.upper())
            value_col = get_column_letter(search_col_idx + 1)
            
            # 遍历指定列查找关键词
            for row in range(1, ws.max_row + 1):
                cell_value = ws[f'{search_column.upper()}{row}'].value
                if cell_value and search_keyword in str(cell_value):
                    # 找到后，读取右侧单元格的值
                    settlement_value = ws[f'{value_col}{row}'].value
                    # 尝试转换为数字
                    if settlement_value is not None:
                        try:
                            return float(settlement_value)
                        except (ValueError, TypeError):
                            return settlement_value
                    return None
            return None
        except Exception as e:
            print(f"搜索 {search_keyword} 失败: {e}")
            return None
    
    def extract_data_from_file(self, file_path, mappings, 
                              search_column="D", search_keyword="折后总计"):
        """
        从单个Excel文件中根据映射配置提取数据
        
        Args:
            file_path: Excel文件路径
            mappings: 映射配置列表，每个映射包含 name 和 cell
            search_column: 搜索结算金额的列
            search_keyword: 搜索的关键词
        
        Returns:
            提取的数据字典
        """
        data = {"文件名": os.path.basename(file_path)}
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            for mapping in mappings:
                cell_ref = mapping['cell'].upper()
                try:
                    value = ws[cell_ref].value
                    # 处理日期格式
                    if isinstance(value, datetime):
                        value = value.strftime("%Y-%m-%d")
                    data[mapping['name']] = value
                except Exception as e:
                    print(f"读取单元格 {cell_ref} 失败: {e}")
                    data[mapping['name']] = None
            
            # 自动搜索并提取结算金额
            settlement_amount = self.find_settlement_amount(ws, search_column, search_keyword)
            data["结算金额"] = settlement_amount
            
            wb.close()
            return data
            
        except Exception as e:
            print(f"处理文件 {file_path} 失败: {e}")
            return None
    
    def merge_bills(self, file_list, mappings, output_file, 
                   search_column="D", search_keyword="折后总计"):
        """
        合并多个账单文件
        
        Args:
            file_list: 要合并的文件路径列表
            mappings: 映射配置列表
            output_file: 输出文件路径
            search_column: 搜索结算金额的列
            search_keyword: 搜索的关键词
        
        Returns:
            处理结果字典
        """
        result = {
            "success": False,
            "success_count": 0,
            "error_count": 0,
            "message": "",
            "data": []
        }
        
        try:
            # 创建新的工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "合并结果"
            
            # 写入表头（添加"序号"列作为第一列，以及"结算金额"列）
            headers = ["序号", "文件名"] + [m['name'] for m in mappings] + ["结算金额"]
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)
                # 设置表头样式
                cell = ws.cell(row=1, column=col_idx)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(
                    start_color="CCE5FF",
                    end_color="CCE5FF",
                    fill_type="solid"
                )
            
            # 处理每个文件
            current_row = 2
            serial_number = 1  # 递增序号
            for file_path in file_list:
                try:
                    data = self.extract_data_from_file(
                        file_path, mappings, search_column, search_keyword
                    )
                    
                    if data:
                        # 写入数据行
                        for col_idx, header in enumerate(headers, start=1):
                            if header == "序号":
                                # 第一列写入递增序号
                                ws.cell(row=current_row, column=col_idx, value=serial_number)
                            else:
                                value = data.get(header)
                                ws.cell(row=current_row, column=col_idx, value=value)
                        
                        serial_number += 1  # 序号递增
                        current_row += 1
                        result["success_count"] += 1
                        result["data"].append(data)
                    else:
                        result["error_count"] += 1
                        
                except Exception as e:
                    print(f"处理文件 {file_path} 时出错: {e}")
                    result["error_count"] += 1
            
            # 自动调整列宽
            for col_idx in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
            
            # 保存结果
            wb.save(output_file)
            wb.close()
            
            result["success"] = True
            result["message"] = "合并完成"
            
        except Exception as e:
            result["success"] = False
            result["message"] = str(e)
        
        return result
    
    def preview_file(self, file_path, max_rows=10, max_cols=10):
        """
        预览Excel文件内容
        
        Args:
            file_path: Excel文件路径
            max_rows: 最大行数
            max_cols: 最大列数
        
        Returns:
            二维数组表示的预览数据
        """
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            
            preview_data = []
            for row_idx in range(1, min(ws.max_row + 1, max_rows + 1)):
                row_data = []
                for col_idx in range(1, min(ws.max_column + 1, max_cols + 1)):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    if isinstance(value, datetime):
                        value = value.strftime("%Y-%m-%d")
                    row_data.append(value if value is not None else "")
                preview_data.append(row_data)
            
            wb.close()
            return preview_data
            
        except Exception as e:
            print(f"预览文件 {file_path} 失败: {e}")
            return None
    
    def get_cell_reference(self, row, col):
        """
        将行列索引转换为单元格引用
        
        Args:
            row: 行号（从1开始）
            col: 列号（从1开始）
        
        Returns:
            单元格引用字符串，如 'A1'
        """
        return f"{get_column_letter(col)}{row}"

